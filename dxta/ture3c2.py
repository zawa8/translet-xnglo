from openpyxl import load_workbook
import shutil

def delete_and_shift_up(ws, col, start_row):
    for r in range(start_row, ws.max_row):
        ws.cell(row=r, column=col).value = ws.cell(row=r + 1, column=col).value
    ws.cell(row=ws.max_row, column=col).value = None

# --- Settings ---
input_file = "xwdrs.xlsx"   # path to your file (will be modified directly)
sheet_name = "t"           # None = active sheet, or specify e.g. "Sheet1"
make_backup = True

# --- Optional backup ---
if make_backup:
    shutil.copy(input_file, input_file + ".bak")

wb = load_workbook(input_file)
ws = wb[sheet_name]

starts_with_ture = []
ends_with_ture = []

col = 7
row = 1

while row <= ws.max_row:
    cell = ws.cell(row=row, column=col)
    word = cell.value

    if not word or not isinstance(word, str):
        row += 1
        continue

    w = word.strip().lower()

    if w.startswith("ture"):
        starts_with_ture.append(word)

        delete_and_shift_up(ws, col, row)
        # ❗ DO NOT increment row here

    elif w.endswith("ture"):
        ends_with_ture.append(word)

        delete_and_shift_up(ws, col, row)
        # ❗ stay on same row

    else:
        row += 1  # move only if no deletion


# Write results
for i, word in enumerate(starts_with_ture, start=1):
    ws.cell(row=i, column=8, value=word)

for i, word in enumerate(ends_with_ture, start=1):
    ws.cell(row=i, column=9, value=word)

wb.save("xwdrs.xlsx")