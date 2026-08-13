import openpyxl
import shutil

# --- Settings ---
input_file = "xwdrs.xlsx"   # path to your file (will be modified directly)
sheet_name = "t"           # None = active sheet, or specify e.g. "Sheet1"
make_backup = True

# --- Optional backup ---
if make_backup:
    shutil.copy(input_file, input_file + ".bak")

# --- Load workbook ---
wb = openpyxl.load_workbook(input_file)
ws = wb[sheet_name] if sheet_name else wb.active

starts_with_ture = []
ends_with_ture = []

# --- Read column 1 and sort words ---
row = 1
while True:
    cell = ws.cell(row=row, column=7)
    word = cell.value

    if word is None:
        break  # stop at first empty cell

    if isinstance(word, str):
        w = word.strip()
        lw = w.lower()

        if lw.startswith("ture"):
            starts_with_ture.append(w)
            ws.cell(row=row, column=7, value=None)
        elif lw.endswith("ture"):
            ends_with_ture.append(w)
            ws.cell(row=row, column=7, value=None)

    row += 1

# --- Write column 2 (packed from row 1) ---
for i, w in enumerate(starts_with_ture, start=1):
    ws.cell(row=i, column=8, value=w)

# --- Write column 3 (packed from row 1) ---
for i, w in enumerate(ends_with_ture, start=1):
    ws.cell(row=i, column=9, value=w)

# --- Save back to the same file ---
wb.save(input_file)
print(f"Done. {len(starts_with_ture)} words in column 8, {len(ends_with_ture)} words in column 9.")